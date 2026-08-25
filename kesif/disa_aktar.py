#!/usr/bin/env python3
"""Ham kesif kayitlarini zenginlestirip Excel + CSV olarak yazar."""
import json, sys, os, csv
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from eposta_zenginlestir import toplu

HAM = "/private/tmp/claude-501/-Users-onursuay-Desktop-Agency-Wizard/43525271-34a1-4a87-9594-e34cdccb80eb/scratchpad/ankara_ham.json"
HEDEF_DIZIN = os.path.expanduser("~/Desktop/Ankara Güzellik Merkezleri - Lead Listesi")

SUTUNLAR = ["ad","ilce","kategori","adres","telefon","web","instagram",
            "eposta","eposta_tipi","tum_epostalar","elenen_ucuncu_taraf",
            "puan","yorum_sayisi","harita"]
BASLIK = {"ad":"Firma Adı","ilce":"İlçe","kategori":"Kategori","adres":"Adres",
          "telefon":"Telefon","web":"Web Sitesi","instagram":"Instagram",
          "eposta":"E-posta","eposta_tipi":"E-posta Tipi","tum_epostalar":"Tüm E-postalar",
          "elenen_ucuncu_taraf":"Elenen (üçüncü taraf)","puan":"Puan",
          "yorum_sayisi":"Yorum Sayısı","harita":"Harita Bağlantısı"}


def main():
    ham = json.load(open(HAM))
    kayitlar = list(ham.values())
    print(f"{len(kayitlar)} işletme zenginleştiriliyor...", flush=True)
    satirlar = toplu(kayitlar)

    # Once e-postasi olanlar, sonra ilce/ad sirasi
    oncelik = {"kurumsal": 0, "kişisel": 1, "tahmin": 2, "": 3}
    satirlar.sort(key=lambda s: (oncelik.get(s["eposta_tipi"], 3), s["ilce"], s["ad"]))

    os.makedirs(HEDEF_DIZIN, exist_ok=True)
    csv_yol = os.path.join(HEDEF_DIZIN, "ankara_guzellik_merkezleri.csv")
    with open(csv_yol, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([BASLIK[c] for c in SUTUNLAR])
        for s in satirlar:
            w.writerow([s.get(c, "") for c in SUTUNLAR])

    xlsx_yol = ""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title = "Lead Listesi"
        ws.append([BASLIK[c] for c in SUTUNLAR])
        for h in ws[1]:
            h.font = Font(bold=True, color="FFFFFF")
            h.fill = PatternFill("solid", fgColor="1F3A5F")
            h.alignment = Alignment(vertical="center")
        for s in satirlar:
            ws.append([s.get(c, "") for c in SUTUNLAR])
        genislik = {"ad":34,"ilce":13,"kategori":22,"adres":46,"telefon":16,"web":30,
                    "instagram":30,"eposta":32,"eposta_tipi":13,"tum_epostalar":34,
                    "elenen_ucuncu_taraf":26,"puan":7,"yorum_sayisi":13,"harita":16}
        for i, c in enumerate(SUTUNLAR, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = genislik[c]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        xlsx_yol = os.path.join(HEDEF_DIZIN, "ankara_guzellik_merkezleri.xlsx")
        wb.save(xlsx_yol)
    except ImportError:
        print("openpyxl yok, yalnız CSV yazıldı.")

    say = {}
    for s in satirlar:
        say[s["eposta_tipi"] or "yok"] = say.get(s["eposta_tipi"] or "yok", 0) + 1
    elenen = sum(1 for s in satirlar if s["elenen_ucuncu_taraf"])
    print(f"\nTOPLAM: {len(satirlar)} işletme")
    for k in ("kurumsal","kişisel","tahmin","yok"):
        if k in say: print(f"  {k}: {say[k]}")
    print(f"  üçüncü taraf adresi elenen kayıt: {elenen}")
    print(f"  Instagram'ı olup e-postası olmayan: "
          f"{sum(1 for s in satirlar if not s['eposta'] and s['instagram'])}")
    print(f"\nCSV : {csv_yol}")
    if xlsx_yol: print(f"XLSX: {xlsx_yol}")


if __name__ == "__main__":
    main()
