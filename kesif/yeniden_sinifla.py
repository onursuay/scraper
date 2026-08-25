#!/usr/bin/env python3
"""Mevcut CSV'deki aday e-postalari yeni sahiplik kuraliyla yeniden siniflandirir."""
import csv, os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from eposta_zenginlestir import sahiplik, temizle
sys.path.insert(0, "/Users/onursuay/Desktop/Agency Wizard/Google Map Scrapper")
from utils.filters import extract_domain_from_url

DIZIN = os.path.expanduser("~/Desktop/Ankara Güzellik Merkezleri - Lead Listesi")
CSV_YOL = os.path.join(DIZIN, "ankara_guzellik_merkezleri.csv")

satirlar = list(csv.DictReader(open(CSV_YOL, encoding="utf-8-sig")))
alanlar = list(satirlar[0].keys())
kazanilan = 0

for s in satirlar:
    ad, site = s["Firma Adı"], s["Web Sitesi"]
    sdom = extract_domain_from_url(site) if site else ""
    adaylar = set()
    for sutun in ("Tüm E-postalar", "Elenen (üçüncü taraf)"):
        for e in (s.get(sutun) or "").split(","):
            e = temizle(e)
            if e and "@" in e:
                adaylar.add(e)
    if not adaylar:
        continue
    kurumsal = sorted(e for e in adaylar if sahiplik(e, sdom, ad) == "kurumsal")
    kisisel = sorted(e for e in adaylar if sahiplik(e, sdom, ad) == "kisisel")
    elenen = sorted(e for e in adaylar if sahiplik(e, sdom, ad) == "ucuncu_taraf")
    onceki = s["E-posta Tipi"]
    if kurumsal:
        s["E-posta"], s["E-posta Tipi"] = kurumsal[0], "kurumsal"
    elif kisisel:
        s["E-posta"], s["E-posta Tipi"] = kisisel[0], "kişisel"
    s["Tüm E-postalar"] = ", ".join(kurumsal + kisisel)
    s["Elenen (üçüncü taraf)"] = ", ".join(elenen)
    if onceki not in ("kurumsal", "kişisel") and s["E-posta Tipi"] in ("kurumsal", "kişisel"):
        kazanilan += 1
        print(f"  geri kazanıldı: {ad[:36]:38s} -> {s['E-posta']}")

oncelik = {"kurumsal": 0, "kişisel": 1, "tahmin": 2, "": 3}
satirlar.sort(key=lambda s: (oncelik.get(s["E-posta Tipi"], 3), s["İlçe"], s["Firma Adı"]))

with open(CSV_YOL, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=alanlar); w.writeheader(); w.writerows(satirlar)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
wb = Workbook(); ws = wb.active; ws.title = "Lead Listesi"
ws.append(alanlar)
for h in ws[1]:
    h.font = Font(bold=True, color="FFFFFF"); h.fill = PatternFill("solid", fgColor="1F3A5F")
for s in satirlar: ws.append([s.get(c, "") for c in alanlar])
for i, c in enumerate(alanlar, 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = min(46, max(12, len(c) + 8))
ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
wb.save(os.path.join(DIZIN, "ankara_guzellik_merkezleri.xlsx"))

say = {}
for s in satirlar: say[s["E-posta Tipi"] or "yok"] = say.get(s["E-posta Tipi"] or "yok", 0) + 1
print(f"\ngeri kazanılan lead: {kazanilan}")
print("dağılım:", {k: say.get(k, 0) for k in ("kurumsal", "kişisel", "tahmin", "yok")})
print("gerçek (kurumsal+kişisel):", say.get("kurumsal", 0) + say.get("kişisel", 0))
