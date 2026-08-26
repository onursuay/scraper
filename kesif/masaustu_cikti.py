#!/usr/bin/env python3
"""Scanner sayfasindaki guncel veriyi masaustune Excel + CSV olarak yazar.

Tek dogruluk kaynagi TABLODUR - bu betik oradan okur, yerel dosyayi degil.
"""
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

from config import SERVICE_ACCOUNT_FILE, GOOGLE_SHEET_URL, SHEET_NAME

SID = GOOGLE_SHEET_URL.split("/d/")[1].split("/")[0]
DIZIN = os.path.expanduser("~/Desktop/Ankara Güzellik Merkezleri - Lead Listesi")
ONCELIK = {"kurumsal": 0, "kişisel": 1, "tahmin": 2, "": 3}


def main():
    c = Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    s = build("sheets", "v4", credentials=c, cache_discovery=False)
    v = s.spreadsheets().values().get(
        spreadsheetId=SID, range=f"'{SHEET_NAME}'!A1:W2000"
    ).execute().get("values", [])
    baslik, satirlar = v[0], [r for r in v[1:] if r and r[0].strip()]
    genislik = len(baslik)
    satirlar = [r + [""] * (genislik - len(r)) for r in satirlar]

    ti = baslik.index("Tip")
    ii = baslik.index("İlçe")
    ai = baslik.index("Firma Adı")
    satirlar.sort(key=lambda r: (ONCELIK.get(r[ti], 3), r[ii], r[ai]))

    os.makedirs(DIZIN, exist_ok=True)
    csv_yol = os.path.join(DIZIN, "ankara_guzellik_merkezleri.csv")
    with open(csv_yol, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(baslik)
        w.writerows(satirlar)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Listesi"
    ws.append(baslik)
    for h in ws[1]:
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="1F3A5F")
    for r in satirlar:
        ws.append(r)
    for i, b in enumerate(baslik, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = min(46, max(12, len(b) + 10))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    xlsx_yol = os.path.join(DIZIN, "ankara_guzellik_merkezleri.xlsx")
    wb.save(xlsx_yol)

    # --- diskten geri okuyarak dogrula ---
    geri = list(csv.reader(open(csv_yol, encoding="utf-8-sig")))
    print(f"tablodan okunan satır : {len(satirlar)}")
    print(f"CSV'ye yazılan satır  : {len(geri) - 1}")
    assert len(geri) - 1 == len(satirlar), "UYUŞMAZLIK"

    say = {}
    for r in satirlar:
        say[r[ti] or "yok"] = say.get(r[ti] or "yok", 0) + 1
    print("tip dağılımı:", {k: say.get(k, 0) for k in ("kurumsal", "kişisel", "tahmin", "yok")})
    if "E-posta Doğrulama" in baslik:
        di = baslik.index("E-posta Doğrulama")
        d = {}
        for r in satirlar:
            if r[di]:
                k = r[di].split(" (")[0]
                d[k] = d.get(k, 0) + 1
        print("doğrulama:", d)
    print(f"\nCSV : {csv_yol}\nXLSX: {xlsx_yol}")


if __name__ == "__main__":
    main()
